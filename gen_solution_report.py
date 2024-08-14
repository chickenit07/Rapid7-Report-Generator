import pandas as pd
import lxml.etree as ET
import sys
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from tqdm import tqdm
import os

def parse_xml_for_solutions_and_products(xml_file, vuln_ids):
    tree = ET.parse(xml_file)
    root = tree.getroot()
    solutions = {}
    ip_product_map = {}

    # Extract product names for each IP address
    for node in root.xpath(".//node"):
        ip_address = node.get("address")
        fingerprint = node.find(".//fingerprints/os")
        if fingerprint is not None:
            product_name = fingerprint.get("product")
            ip_product_map[ip_address] = product_name

    for vuln_id in tqdm(vuln_ids, desc="Parsing XML for Solutions"):
        xpath_query = f".//vulnerability[@id='{vuln_id}']/solution/ContainerBlockElement/UnorderedList/ListItem"
        list_items = root.xpath(xpath_query)
        
        solution_list = []
        for item in list_items:
            paragraphs = item.xpath(".//Paragraph")
            item_texts = []

            for i, para in enumerate(paragraphs):
                text_parts = []
                for elem in para.iter():
                    if elem.tag == "URLLink":
                        text_parts.append(elem.get("LinkURL"))
                    elif elem.text:
                        text_parts.append(elem.text.strip())
                paragraph_text = "".join(text_parts)

                if i == 0:
                    item_texts.append(paragraph_text)
                else:
                    item_texts.append(f"=> {paragraph_text}")
            
            full_item_text = " ".join(item_texts)
            solution_list.append(full_item_text)

        solutions[vuln_id] = solution_list

    return solutions, ip_product_map

def read_and_sort_csv(csv_file):
    print("Reading and sorting CSV file...")
    df = pd.read_csv(csv_file)
    df_sorted = df.sort_values(by="Asset IP Address")
    print(f"CSV file loaded with {len(df)} records.")
    return df_sorted

def process_vulnerabilities(df_sorted, solutions, ip_product_map):
    print("Processing vulnerabilities...")
    output_data_main = []
    output_data_windows = []

    for _, row in tqdm(df_sorted.iterrows(), total=df_sorted.shape[0], desc="Processing Rows"):
        ip_address = row['Asset IP Address']
        vuln_id = row['Vulnerability ID']
        solution_items = solutions.get(vuln_id, [])

        if vuln_id.startswith("msft"):
            product_name = ip_product_map.get(ip_address, "")
            filtered_solutions = [sol for sol in solution_items if product_name in sol]
            
            for solution in filtered_solutions:
                # Add product name to IP address
                ip_with_product = f"{ip_address} - {product_name}" if product_name else ip_address
                output_data_windows.append([ip_with_product, solution])
        else:
            for solution in solution_items:
                output_data_main.append([ip_address, solution])

    return output_data_main, output_data_windows

def create_dataframes(output_data_main, output_data_windows):
    print("Converting processed data to DataFrames...")
    df_main = pd.DataFrame(output_data_main, columns=['Asset IP Address', 'Solution'])
    df_windows = pd.DataFrame(output_data_windows, columns=['Asset IP Address', 'Solution'])
    return df_main, df_windows

def process_linux_dataframe(df_main):
    print("Processing solutions for Linux...")
    df_main['Services'] = df_main['Solution'].apply(lambda x: x.split('=>')[0].strip())
    df_main['Solution Details'] = df_main['Solution'].apply(lambda x: '=>'.join(x.split('=>')[1:]).strip())
    df_main.drop(columns=['Solution'], inplace=True)

    df_main = df_main.drop_duplicates(subset=['Asset IP Address', 'Services']).copy()

    # Ensure the DataFrame has a proper index
    df_main.reset_index(drop=True, inplace=True)
    
    # Add Owner column with a blank space
    df_main['Owner'] = ' '

    return df_main

def process_windows_dataframe(df_windows):
    print("Splitting solutions for Windows...")
    df_windows['Services'] = df_windows['Solution'].apply(lambda x: '=>'.join(x.split('=>')[:-1]).strip())
    df_windows['Solution Details'] = df_windows['Solution'].apply(lambda x: x.split('=>')[-1].strip())
    df_windows.drop(columns=['Solution'], inplace=True)

    df_windows = df_windows.copy()

    # Ensure the DataFrame has a proper index
    df_windows.reset_index(drop=True, inplace=True)
    
    # Add Owner column with a blank space
    df_windows['Owner'] = ' '

    return df_windows


def save_to_excel(df_main, df_windows, output_file):
    print("Writing data to Excel file...")
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_main.to_excel(writer, index=False, sheet_name='Linux')
        df_windows.to_excel(writer, index=False, sheet_name='Windows')
    
    print("Setting column widths and formatting...")
    wb = load_workbook(output_file)
    linux_sheet = wb['Linux']
    windows_sheet = wb['Windows']

    # Set the column widths
    for col, width in zip(['A', 'B', 'C', 'D'], [30, 80, 100, 30]):
        linux_sheet.column_dimensions[col].width = width
        windows_sheet.column_dimensions[col].width = width

    def merge_cells(sheet, col):
        current_value = None
        start_row = 2
        
        for row in range(2, sheet.max_row + 1):
            cell = sheet[f'{col}{row}']
            if cell.value != current_value:
                if start_row < row - 1:
                    sheet.merge_cells(start_row=start_row, start_column=1, end_row=row - 1, end_column=1)
                current_value = cell.value
                start_row = row
        
        if start_row < sheet.max_row:
            sheet.merge_cells(start_row=start_row, start_column=1, end_row=sheet.max_row, end_column=1)

    def set_alignment(sheet):
        for cell in sheet['A']:
            cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

    merge_cells(linux_sheet, 'A')
    merge_cells(windows_sheet, 'A')

    set_alignment(linux_sheet)
    set_alignment(windows_sheet)
    
    wb.save(output_file)
    print(f"Output successfully written to {output_file}")

def main(csv_file, xml_file):
    print("Starting script...")
    
    df_sorted = read_and_sort_csv(csv_file)
    unique_vuln_ids = df_sorted['Vulnerability ID'].unique()
    solutions, ip_product_map = parse_xml_for_solutions_and_products(xml_file, unique_vuln_ids)
    
    output_data_main, output_data_windows = process_vulnerabilities(df_sorted, solutions, ip_product_map)
    df_main, df_windows = create_dataframes(output_data_main, output_data_windows)

    df_main = process_linux_dataframe(df_main)
    df_windows = process_windows_dataframe(df_windows)

    # Generate output file name based on CSV file name
    base_name = os.path.splitext(os.path.basename(csv_file))[0]
    output_file = os.path.join(os.getcwd(), f"{base_name}_Solution_Details.xlsx")
    
    save_to_excel(df_main, df_windows, output_file)

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: script.py <csv_file> <xml_file>")
        sys.exit(1)
    
    csv_file = sys.argv[1]
    xml_file = sys.argv[2]

    main(csv_file, xml_file)
